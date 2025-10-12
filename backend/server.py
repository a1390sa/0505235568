from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
from datetime import datetime, timedelta
from bson import ObjectId
import openpyxl
import io
from hijri_converter import Hijri, Gregorian


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# Define Models
class Task(BaseModel):
    id: Optional[str] = None
    userId: str
    field: str  # المجال
    name: str  # المهمة
    description: str = ""  # الوصف
    implementation_method: str = ""  # آلية التنفيذ
    work_type: str = "office"  # مكتبي/ميداني (office/field)
    duration_days: int = 1  # عدد الأيام
    start_date: str  # تاريخ البداية (ميلادي)
    start_date_hijri: Optional[str] = None  # تاريخ البداية (هجري)
    recurrence_type: str = "none"  # none, daily, weekly, monthly
    recurrence_end_date: Optional[str] = None  # تاريخ نهاية التكرار
    work_days_only: bool = True  # فقط أيام العمل (الأحد-الخميس)
    priority: str = "medium"  # "high", "medium", "low"
    completed: bool = False
    createdAt: Optional[datetime] = None
    updatedAt: Optional[datetime] = None

class TaskCreate(BaseModel):
    userId: str
    field: str
    name: str
    description: str = ""
    implementation_method: str = ""
    work_type: str = "office"
    duration_days: int = 1
    start_date: str
    priority: str = "medium"

class TaskUpdate(BaseModel):
    field: Optional[str] = None
    name: Optional[str] = None
    description: Optional[str] = None
    implementation_method: Optional[str] = None
    work_type: Optional[str] = None
    duration_days: Optional[int] = None
    start_date: Optional[str] = None
    priority: Optional[str] = None
    completed: Optional[bool] = None


# Helper function to convert MongoDB document to Task
def task_helper(task) -> dict:
    return {
        "id": str(task["_id"]),
        "userId": task.get("userId"),
        "field": task.get("field"),
        "name": task.get("name"),
        "description": task.get("description", ""),
        "implementation_method": task.get("implementation_method", ""),
        "work_type": task.get("work_type", "office"),
        "duration_days": task.get("duration_days", 1),
        "start_date": task.get("start_date"),
        "priority": task.get("priority", "medium"),
        "completed": task.get("completed", False),
        "createdAt": task.get("createdAt"),
        "updatedAt": task.get("updatedAt"),
    }


# Routes
@api_router.get("/")
async def root():
    return {"message": "Task Manager API"}


@api_router.post("/tasks", response_model=dict)
async def create_task(task: TaskCreate):
    task_dict = task.dict()
    task_dict["createdAt"] = datetime.utcnow()
    task_dict["updatedAt"] = datetime.utcnow()
    task_dict["completed"] = False
    
    result = await db.tasks.insert_one(task_dict)
    new_task = await db.tasks.find_one({"_id": result.inserted_id})
    
    return task_helper(new_task)


@api_router.get("/tasks", response_model=List[dict])
async def get_tasks(
    userId: str,
    work_type: Optional[str] = None,
    field: Optional[str] = None,
    completed: Optional[bool] = None,
    priority: Optional[str] = None
):
    query = {"userId": userId}
    
    if work_type:
        query["work_type"] = work_type
    if field:
        query["field"] = field
    if completed is not None:
        query["completed"] = completed
    if priority:
        query["priority"] = priority
    
    tasks = await db.tasks.find(query).sort("start_date", 1).to_list(1000)
    return [task_helper(task) for task in tasks]


@api_router.get("/tasks/{task_id}", response_model=dict)
async def get_task(task_id: str):
    try:
        task = await db.tasks.find_one({"_id": ObjectId(task_id)})
        if task:
            return task_helper(task)
        raise HTTPException(status_code=404, detail="Task not found")
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.put("/tasks/{task_id}", response_model=dict)
async def update_task(task_id: str, task_update: TaskUpdate):
    try:
        update_data = {k: v for k, v in task_update.dict().items() if v is not None}
        
        if not update_data:
            raise HTTPException(status_code=400, detail="No data to update")
        
        update_data["updatedAt"] = datetime.utcnow()
        
        result = await db.tasks.update_one(
            {"_id": ObjectId(task_id)},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Task not found")
        
        updated_task = await db.tasks.find_one({"_id": ObjectId(task_id)})
        return task_helper(updated_task)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.delete("/tasks/{task_id}")
async def delete_task(task_id: str):
    try:
        result = await db.tasks.delete_one({"_id": ObjectId(task_id)})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Task not found")
        
        return {"message": "Task deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.post("/tasks/import")
async def import_tasks(userId: str, file: UploadFile = File(...)):
    try:
        # Read the uploaded file
        contents = await file.read()
        
        # Load the Excel file
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active
        
        tasks_imported = 0
        errors = []
        
        # Find column indices from header row
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        # Map column names to indices - case insensitive and flexible matching
        col_map = {}
        for idx, header in enumerate(header_row):
            if header:
                header_str = str(header).strip()
                # Check if contains Arabic text
                if "مجال" in header_str:
                    col_map["field"] = idx
                if "مهمة" in header_str or "برنامج" in header_str:
                    col_map["name"] = idx
                if "وصف" in header_str:
                    col_map["description"] = idx
                if "آلية" in header_str:
                    col_map["implementation_method"] = idx
                if "مكتبي" in header_str or "ميداني" in header_str:
                    col_map["work_type"] = idx
                if "أيام" in header_str or ("عدد" in header_str and idx not in col_map.values()):
                    col_map["duration_days"] = idx
                if "تاريخ" in header_str and "بداية" in header_str:
                    col_map["start_date"] = idx
        
        # Log found columns for debugging
        logger.info(f"Detected columns: {col_map}")
        logger.info(f"Header row: {header_row}")
        
        # Process data rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Get name field (required)
                name_idx = col_map.get("name", 12)  # Default to column 12
                task_name = str(row[name_idx]).strip() if len(row) > name_idx and row[name_idx] else ""
                
                # Skip empty rows or rows with very short names (likely numbers or invalid data)
                if not task_name or len(task_name) < 3 or task_name.isdigit():
                    continue
                
                # Get other fields with safe defaults
                field_idx = col_map.get("field", 11)
                desc_idx = col_map.get("description", 13)
                impl_idx = col_map.get("implementation_method", 14)
                work_idx = col_map.get("work_type", 15)
                days_idx = col_map.get("duration_days", 16)
                date_idx = col_map.get("start_date", 17)
                
                # Parse work_type
                work_type_raw = str(row[work_idx]).strip() if len(row) > work_idx and row[work_idx] else "مكتبي"
                work_type = "field" if "ميداني" in work_type_raw else "office"
                
                # Parse duration_days
                duration_days = 1
                if len(row) > days_idx and row[days_idx]:
                    days_str = str(row[days_idx])
                    # Extract numbers from string like "4 أيام" or "يوم واحد"
                    import re
                    numbers = re.findall(r'\d+', days_str)
                    if numbers:
                        duration_days = int(numbers[0])
                    elif "واحد" in days_str:
                        duration_days = 1
                
                # Parse start_date
                start_date_val = datetime.now().date().isoformat()
                if len(row) > date_idx and row[date_idx]:
                    date_str = str(row[date_idx])
                    # Handle both formats: 1447-4-29 and 2025-07-20
                    try:
                        if isinstance(row[date_idx], datetime):
                            start_date_val = row[date_idx].date().isoformat()
                        else:
                            start_date_val = date_str
                    except:
                        pass
                
                task_dict = {
                    "userId": userId,
                    "field": str(row[field_idx]).strip() if len(row) > field_idx and row[field_idx] else "عام",
                    "name": task_name,
                    "description": str(row[desc_idx]).strip() if len(row) > desc_idx and row[desc_idx] else "",
                    "implementation_method": str(row[impl_idx]).strip() if len(row) > impl_idx and row[impl_idx] else "",
                    "work_type": work_type,
                    "duration_days": duration_days,
                    "start_date": start_date_val,
                    "priority": "medium",
                    "completed": False,
                    "createdAt": datetime.utcnow(),
                    "updatedAt": datetime.utcnow(),
                }
                
                await db.tasks.insert_one(task_dict)
                tasks_imported += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        return {
            "message": "Import completed",
            "tasks_imported": tasks_imported,
            "errors": errors if errors else None
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to import file: {str(e)}")


@api_router.get("/tasks/upcoming")
async def get_upcoming_tasks(userId: str):
    """Get upcoming tasks starting soon"""
    today = datetime.now().date().isoformat()
    
    # Get tasks starting today or later that are not completed
    tasks = await db.tasks.find({
        "userId": userId,
        "start_date": {"$gte": today},
        "completed": False
    }).sort("start_date", 1).to_list(100)
    
    return [task_helper(task) for task in tasks]


@api_router.get("/download/template")
async def download_template():
    """Download Excel template file"""
    from fastapi.responses import FileResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Create template file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "نموذج المهام"
    
    headers = ["المجال", "المهمة", "الوصف", "آلية التنفيذ", "مكتبي/ميداني", "عدد الأيام", "تاريخ البداية"]
    
    header_fill = PatternFill(start_color="2E7D8F", end_color="2E7D8F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    examples = [
        ["التعليم والتدريب", "ورشة تدريبية للموظفين", "ورشة عن مهارات التواصل الفعال", "تدريب حضوري مع مدرب معتمد", "ميداني", 3, "2025-07-15"],
        ["التسويق", "حملة إعلانية على وسائل التواصل", "حملة تسويقية للمنتج الجديد", "تصميم محتوى ونشره على المنصات", "مكتبي", 14, "2025-07-20"],
        ["الموارد البشرية", "اجتماع فريق العمل الشهري", "مناقشة الإنجازات والخطط القادمة", "اجتماع افتراضي عبر Zoom", "مكتبي", 1, "2025-07-25"],
    ]
    
    for row_num, example in enumerate(examples, 2):
        for col_num, value in enumerate(example, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='right' if col_num < 7 else 'center', vertical='center')
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18
    
    template_path = "/tmp/template_download.xlsx"
    wb.save(template_path)
    
    return FileResponse(
        template_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="نموذج_استيراد_المهام.xlsx"
    )


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()